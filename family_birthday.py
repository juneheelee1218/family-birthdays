# 가족생일.py
from datetime import date
from openpyxl import Workbook
from korean_lunar_calendar import KoreanLunarCalendar

LEAP_MONTHS = {2001:4,2004:2,2006:7,2009:5,2012:3,2014:9,2017:5,2020:4,2023:2,2025:6,2028:5,2031:3,2033:11,2036:6,2039:5,2042:2,2044:7,2047:5,2050:3}
WD = ["월","화","수","목","금","토","일"]
def fmt(y,m,d): return f"{m:02d}/{d:02d} ({WD[date(y,m,d).weekday()]})"

def lunar2solar(y, m, d):
    # Try 평달 먼저, then (if applicable) 윤달; if day invalid (e.g., 30), fallback to d-1
    for dd in (d, d-1):  # <- 핵심: 30일 없는 달이면 29일로 자동 처리
        if dd <= 0:
            continue
        for inter in (False, True):
            if inter and LEAP_MONTHS.get(y) != m:
                continue
            try:
                cal = KoreanLunarCalendar()
                cal.setLunarDate(y, m, dd, inter)
                if cal.solarYear > 0:
                    return cal.solarYear, cal.solarMonth, cal.solarDay
            except:
                pass
    raise ValueError(f"Invalid lunar date: {y}년 음력 {m}월 {d}일")

def main():
    years = list(range(2026, 2035))
    people = [
        ("아빠","l", 5,30), # 30일 없는 해는 자동으로 29일로 처리됨
        ("엄마","l", 10, 16),
        ("이모","l",11, 4),
        ("삼촌","l", 7,29),
        ("이모부","l",8,13),
        ("언니","s",7, 3),
        ("나","s",12,18),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Birthdays"

    ws.cell(1,1,"이름")
    for c, y in enumerate(years, start=2):
        ws.cell(1,c,y)

    for r, (name, kind, mm, dd) in enumerate(people, start=2):
        ws.cell(r,1,name)
        for c, y in enumerate(years, start=2):
            if kind == "l":
                sy, sm, sd = lunar2solar(y, mm, dd)
            else:
                sy, sm, sd = y, mm, dd
            ws.cell(r,c,fmt(sy, sm, sd))

    wb.save("family_birthdays.xlsx")
    print("Saved: family_birthdays.xlsx")

if __name__ == "__main__":
    main()
